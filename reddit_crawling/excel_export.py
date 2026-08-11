# excel_export.py
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _autosize_and_style(ws, wrap_cols=None, url_col_name=None):
    """
    - 열 너비 자동 조정(너무 길면 상한)
    - 헤더 스타일
    - 줄바꿈(wrap text)
    - 필터/고정
    - URL 컬럼 하이퍼링크
    """
    wrap_cols = set(wrap_cols or [])

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F4E79")  # 남색 계열
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    # 필터 + 상단 고정
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # 컬럼명 -> 인덱스 매핑
    col_index = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # 하이퍼링크 적용
    if url_col_name and url_col_name in col_index:
        url_col = col_index[url_col_name]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=url_col)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    # 줄바꿈 적용할 컬럼 인덱스
    wrap_col_indices = {col_index[name] for name in wrap_cols if name in col_index}

    # 열 너비 계산
    max_width_cap = 60  # 너무 길면 보기 힘드니까 상한
    min_width = 10
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            # 매우 긴 본문은 너비 계산에서 과대평가 방지
            if len(s) > 120:
                s = s[:120]
            max_len = max(max_len, len(s))

            # wrap 대상 컬럼은 정렬도 세팅
            if r >= 2 and c in wrap_col_indices:
                ws.cell(row=r, column=c).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        width = max(min_width, min(max_width_cap, max_len + 2))
        ws.column_dimensions[letter].width = width

    # 행 높이(본문 컬럼이 있으면 조금 키우기)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 45


def reddit_csv_to_excel(csv_path: str, xlsx_path: str) -> str:
    """
    네가 저장한 reddit CSV(댓글 단위 row)를 받아서
    - Posts 시트: 게시글 단위로 중복 제거 + 댓글수
    - Comments 시트: 댓글 단위 그대로
    로 정리해서 xlsx로 저장.
    """
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    df = pd.read_csv(csv_path)

    # 안전하게 컬럼 확인(너 파일 기준)
    needed = [
        "keyword", "subreddit", "relevance_score",
        "post_url", "post_title", "post_body",
        "comment_author", "comment_text"
    ]
    for col in needed:
        if col not in df.columns:
            raise ValueError(f"Missing column in CSV: {col}")

    # Comments 시트(댓글 단위)
    comments = df.copy()

    # Posts 시트(게시글 단위)
    # 게시글을 고유키(post_url)로 묶어서 댓글 수, 대표 keyword/subreddit/score 등을 정리
    posts = (
        df.groupby(["post_url"], as_index=False)
          .agg(
              keyword=("keyword", "first"),
              subreddit=("subreddit", "first"),
              relevance_score=("relevance_score", "max"),
              post_title=("post_title", "first"),
              post_body=("post_body", "first"),
              comment_count=("comment_text", "count"),
          )
    )

    # 보기 좋게 정렬: 점수 높은 순
    posts = posts.sort_values(["relevance_score", "comment_count"], ascending=[False, False])

    # 저장 디렉토리 생성
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    # 엑셀 쓰기
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        posts.to_excel(writer, sheet_name="Posts", index=False)
        comments.to_excel(writer, sheet_name="Comments", index=False)

    # openpyxl로 스타일링
    wb = load_workbook(xlsx_path)

    ws_posts = wb["Posts"]
    _autosize_and_style(
        ws_posts,
        wrap_cols=["post_title", "post_body"],
        url_col_name="post_url",
    )

    ws_comments = wb["Comments"]
    _autosize_and_style(
        ws_comments,
        wrap_cols=["comment_text", "post_title", "post_body"],
        url_col_name="post_url",
    )

    wb.save(xlsx_path)
    return xlsx_path

def main(csv_path=None, xlsx_path=None):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, "data")

    csv_path = csv_path or os.path.join(data_dir, "reddit_top_relevant_month.csv")
    xlsx_path = xlsx_path or os.path.join(data_dir, "reddit_top_relevant_month.xlsx")

    out = reddit_csv_to_excel(csv_path, xlsx_path)
    print(f"✅ Excel 생성 완료: {out}")
    return out


if __name__ == "__main__":
    main()
